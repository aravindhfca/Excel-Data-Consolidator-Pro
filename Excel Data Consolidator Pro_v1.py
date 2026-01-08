import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import csv
import json
import warnings
import webbrowser

# Try to import pyxlsb for .xlsb support
try:
    from pyxlsb import open_workbook as open_xlsb
    XLSB_SUPPORT = True
except ImportError:
    XLSB_SUPPORT = False
    warnings.warn("pyxlsb not installed. .xlsb format support disabled. Install with: pip install pyxlsb")

class ExcelMergerApp:
    def __init__(self, master):
        self.master = master
        master.title("Excel Data Consolidator Pro")
        master.geometry("950x850")
        master.configure(bg="#f0f0f0")
        
        # Enable window resizing with minimum size
        master.minsize(800, 600)
        
        self.filepaths = []
        self.preview_data = []
        self.theme = "light"
        self.config_file = Path.home() / ".excel_merger_config.json"
        self.load_config()
        self.selected_sheets = []
        self.sheet_names_per_file = {}
        self.log_file = Path.home() / "excel_merger_error_log.txt"
        self.available_headers = []
        
        self.setup_ui()
        
    def setup_ui(self):
        style = ttk.Style()
        style.theme_use('clam')
        
        main_frame = tk.Frame(self.master, bg="#f0f0f0")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Header with Designer Info
        header_frame = tk.Frame(main_frame, bg="#2c3e50", height=75)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text="📊 Excel Data Consolidator Pro", 
                               font=("Arial", 16, "bold"), bg="#2c3e50", fg="white")
        title_label.pack(side=tk.LEFT, padx=20, pady=5)
        
        # Designer Info Frame
        info_frame = tk.Frame(header_frame, bg="#34495e")
        info_frame.pack(side=tk.LEFT, padx=20, pady=5)
        
        designer_label = tk.Label(info_frame, text="Designed by: CA Aravindh K", 
                                 font=("Arial", 9, "bold"), bg="#34495e", fg="#ecf0f1")
        designer_label.pack()
        
        github_btn = tk.Button(info_frame, text="🔗 GitHub Profile", font=("Arial", 8),
                              command=self.open_github, bd=0, bg="#1abc9c", fg="white",
                              activebackground="#16a085", cursor="hand2", padx=5, pady=2)
        github_btn.pack()
        
        theme_btn = tk.Button(header_frame, text="🌓", font=("Arial", 14),
                             command=self.toggle_theme, bd=0, bg="#34495e", fg="white",
                             activebackground="#1abc9c", cursor="hand2")
        theme_btn.pack(side=tk.RIGHT, padx=20)
        
        content_frame = tk.Frame(main_frame, bg="#f0f0f0")
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Use PanedWindow for flexible layout
        paned_window = ttk.PanedWindow(content_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True)
        
        left_panel = tk.Frame(paned_window, bg="white", relief=tk.RAISED, bd=1)
        paned_window.add(left_panel, weight=1)
        
        file_section = self.create_section(left_panel, "📁 File Selection")
        
        # Update file types in button
        file_types_text = "Select Excel/CSV Files"
        if XLSB_SUPPORT:
            file_types_text += " (.xlsx, .xlsm, .xlsb, .csv)"
        else:
            file_types_text += " (.xlsx, .xlsm, .csv)"
        
        self.select_btn = tk.Button(file_section, text=file_types_text, 
                                    command=self.select_files, bg="#3498db", fg="white",
                                    font=("Arial", 9, "bold"), cursor="hand2",
                                    activebackground="#2980b9", relief=tk.FLAT, padx=10, pady=8)
        self.select_btn.pack(pady=5, fill=tk.X, padx=10)
        
        self.files_listbox = tk.Listbox(file_section, height=6, bg="#ecf0f1", 
                                       selectmode=tk.MULTIPLE)
        self.files_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Enable drag-drop for files_listbox
        self.setup_drag_drop(self.files_listbox)
        
        btn_frame = tk.Frame(file_section, bg="white")
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Button(btn_frame, text="Clear", command=self.clear_files,
                 bg="#e74c3c", fg="white", relief=tk.FLAT, cursor="hand2", 
                 font=("Arial", 8)).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Remove", command=self.remove_selected,
                 bg="#e67e22", fg="white", relief=tk.FLAT, cursor="hand2",
                 font=("Arial", 8)).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Auto Headers", command=self.auto_detect_headers,
                 bg="#16a085", fg="white", relief=tk.FLAT, cursor="hand2",
                 font=("Arial", 8)).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Select Sheets", command=self.select_sheets_dialog,
                 bg="#f39c12", fg="white", relief=tk.FLAT, cursor="hand2",
                 font=("Arial", 8)).pack(side=tk.LEFT, padx=2)
        
        config_section = self.create_section(left_panel, "⚙️ Configuration")
        
        # Sheet Name with Dropdown
        tk.Label(config_section, text="Sheet Name (optional):", bg="white", 
                font=("Arial", 9)).grid(row=0, column=0, sticky="w", padx=10, pady=3)
        self.sheet_name_var = tk.StringVar()
        self.sheet_name_combo = ttk.Combobox(config_section, textvariable=self.sheet_name_var, 
                                            width=20, state="readonly", font=("Arial", 9))
        self.sheet_name_combo.grid(row=0, column=1, sticky="ew", padx=5, pady=3)
        config_section.columnconfigure(1, weight=1)
        
        # Header Dropdown
        tk.Label(config_section, text="Select Header Row:", bg="white", 
                font=("Arial", 9)).grid(row=1, column=0, sticky="w", padx=10, pady=3)
        self.header_var = tk.StringVar(value="first")
        self.header_combo = ttk.Combobox(config_section, textvariable=self.header_var,
                                        values=["Skip First", "Include All", "First Only"],
                                        width=20, state="readonly", font=("Arial", 9))
        self.header_combo.grid(row=1, column=1, sticky="ew", padx=5, pady=3)
        
        # Start Row
        self.create_input_row(config_section, "Start Row:", "row_entry", "1", row=2)
        
        # Start Column
        self.create_input_row(config_section, "Start Column:", "col_entry", "A", row=3)
        
        # Column Headers Listbox
        tk.Label(config_section, text="Column Headers:", bg="white", 
                font=("Arial", 9, "bold")).grid(row=4, column=0, sticky="w", padx=10, pady=(10, 3))
        
        headers_frame = tk.Frame(config_section, bg="white")
        headers_frame.grid(row=5, column=0, columnspan=2, sticky="ew", padx=10, pady=3)
        
        scroll_headers = tk.Scrollbar(headers_frame, orient=tk.VERTICAL)
        scroll_headers.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.headers_listbox = tk.Listbox(headers_frame, height=4, bg="#ecf0f1",
                                         yscrollcommand=scroll_headers.set, selectmode=tk.MULTIPLE)
        self.headers_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_headers.config(command=self.headers_listbox.yview)
        
        headers_btn_frame = tk.Frame(config_section, bg="white")
        headers_btn_frame.grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=3)
        
        tk.Button(headers_btn_frame, text="Detect Headers", command=self.detect_column_headers,
                 bg="#3498db", fg="white", relief=tk.FLAT, cursor="hand2", 
                 font=("Arial", 8)).pack(side=tk.LEFT, padx=2)
        tk.Button(headers_btn_frame, text="Clear Selection", command=self.clear_headers_selection,
                 bg="#95a5a6", fg="white", relief=tk.FLAT, cursor="hand2",
                 font=("Arial", 8)).pack(side=tk.LEFT, padx=2)
        
        options_section = self.create_section(left_panel, "🎯 Options")
        
        self.skip_empty_var = tk.BooleanVar(value=True)
        tk.Checkbutton(options_section, text="Skip empty rows", 
                      variable=self.skip_empty_var, bg="white", font=("Arial", 9)).pack(anchor="w", padx=10)
        
        self.remove_duplicates_var = tk.BooleanVar(value=False)
        tk.Checkbutton(options_section, text="Remove duplicate rows", 
                      variable=self.remove_duplicates_var, bg="white", font=("Arial", 9)).pack(anchor="w", padx=10)
        
        self.add_source_var = tk.BooleanVar(value=False)
        tk.Checkbutton(options_section, text="Add source filename column", 
                      variable=self.add_source_var, bg="white", font=("Arial", 9)).pack(anchor="w", padx=10)
        
        # Sheet Name Column Position
        tk.Label(options_section, text="Sheet Name Column Position:", bg="white", 
                font=("Arial", 9, "bold")).pack(anchor="w", padx=10, pady=(10, 3))
        self.sheet_col_var = tk.StringVar(value="none")
        sheet_col_frame = tk.Frame(options_section, bg="white")
        sheet_col_frame.pack(anchor="w", padx=20, pady=2)
        tk.Radiobutton(sheet_col_frame, text="None", variable=self.sheet_col_var, 
                      value="none", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        tk.Radiobutton(sheet_col_frame, text="First Column", variable=self.sheet_col_var, 
                      value="first", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        tk.Radiobutton(sheet_col_frame, text="Last Column", variable=self.sheet_col_var, 
                      value="last", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        
        # Row ID Option
        self.add_row_id_var = tk.BooleanVar(value=False)
        tk.Checkbutton(options_section, text="Add numerical Row IDs (Column Position):", 
                      variable=self.add_row_id_var, bg="white", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(10, 3))
        self.row_id_var = tk.StringVar(value="none")
        row_id_frame = tk.Frame(options_section, bg="white")
        row_id_frame.pack(anchor="w", padx=20, pady=2)
        tk.Radiobutton(row_id_frame, text="None", variable=self.row_id_var, 
                      value="none", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        tk.Radiobutton(row_id_frame, text="First Column", variable=self.row_id_var, 
                      value="first", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        tk.Radiobutton(row_id_frame, text="Last Column", variable=self.row_id_var, 
                      value="last", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        
        tk.Label(options_section, text="Export Format:", bg="white", 
                font=("Arial", 9, "bold")).pack(anchor="w", padx=10, pady=(10, 2))
        self.format_var = tk.StringVar(value="xlsx")
        format_frame = tk.Frame(options_section, bg="white")
        format_frame.pack(anchor="w", padx=20)
        tk.Radiobutton(format_frame, text="Excel (.xlsx)", variable=self.format_var, 
                      value="xlsx", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        tk.Radiobutton(format_frame, text="CSV (.csv)", variable=self.format_var, 
                      value="csv", bg="white", font=("Arial", 8)).pack(side=tk.LEFT)
        
        action_frame = tk.Frame(left_panel, bg="white", pady=15)
        action_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        self.preview_btn = tk.Button(action_frame, text="👁️ Preview Data", 
                                     command=self.preview_data_action, bg="#9b59b6", fg="white",
                                     font=("Arial", 10, "bold"), cursor="hand2",
                                     relief=tk.FLAT, padx=20, pady=8)
        self.preview_btn.pack(fill=tk.X, padx=10, pady=3)
        
        self.merge_btn = tk.Button(action_frame, text="✓ Merge & Export", 
                                   command=self.merge_files, bg="#27ae60", fg="white",
                                   font=("Arial", 11, "bold"), cursor="hand2",
                                   relief=tk.FLAT, padx=20, pady=10)
        self.merge_btn.pack(fill=tk.X, padx=10, pady=3)
        
        right_panel = tk.Frame(paned_window, bg="white", relief=tk.RAISED, bd=1)
        paned_window.add(right_panel, weight=1)
        
        preview_header = tk.Frame(right_panel, bg="#34495e", height=40)
        preview_header.pack(fill=tk.X)
        preview_header.pack_propagate(False)
        
        tk.Label(preview_header, text="📋 Data Preview", font=("Arial", 12, "bold"),
                bg="#34495e", fg="white").pack(side=tk.LEFT, padx=15, pady=8)
        
        self.row_count_label = tk.Label(preview_header, text="Rows: 0", 
                                       bg="#34495e", fg="#ecf0f1", font=("Arial", 9))
        self.row_count_label.pack(side=tk.RIGHT, padx=15)
        
        preview_frame = tk.Frame(right_panel)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        tree_scroll_y = tk.Scrollbar(preview_frame)
        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree_scroll_x = tk.Scrollbar(preview_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.preview_tree = ttk.Treeview(preview_frame, 
                                        yscrollcommand=tree_scroll_y.set,
                                        xscrollcommand=tree_scroll_x.set,
                                        height=20)
        self.preview_tree.pack(fill=tk.BOTH, expand=True)
        
        tree_scroll_y.config(command=self.preview_tree.yview)
        tree_scroll_x.config(command=self.preview_tree.xview)
        
        self.status_bar = tk.Label(main_frame, text="Ready", bg="#34495e", fg="white",
                                  anchor="w", font=("Arial", 9), relief=tk.SUNKEN)
        self.status_bar.pack(fill=tk.X, pady=(10, 0))
        
        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.pack(fill=tk.X, pady=(5, 0))
        self.progress.pack_forget()
    
    def setup_drag_drop(self, widget):
        """Enable drag-drop file support for the given widget using tkinterdnd2 if available"""
        try:
            from tkinterdnd2 import DND_FILES, DND_TEXT
            
            def drop(event):
                files = self.parse_drag_drop_data(event.data)
                for filepath in files:
                    if filepath not in self.filepaths:
                        self.filepaths.append(filepath)
                self.update_file_list()
                self.update_sheet_dropdown()
                self.update_status(f"Added {len(files)} file(s) via drag-drop")
                return event.action
            
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind('<<Drop>>', drop)
        except ImportError:
            # tkinterdnd2 not available, drag-drop won't work
            pass
    
    def parse_drag_drop_data(self, data):
        """Parse drag-drop data to extract file paths"""
        files = []
        # Handle Windows path format with curly braces: {path1} {path2}
        paths = data.replace('{', '').replace('}', '').split()
        for path in paths:
            path = path.strip()
            if path and Path(path).exists():
                files.append(path)
        return files
    
    def create_section(self, parent, title):
        frame = tk.LabelFrame(parent, text=title, bg="white", 
                             font=("Arial", 10, "bold"), fg="#2c3e50",
                             relief=tk.GROOVE, bd=2)
        frame.pack(fill=tk.BOTH, padx=10, pady=8)
        return frame
    
    def create_input_row(self, parent, label_text, attr_name, default_value, row=None):
        if row is None:
            row = parent.grid_size()[1]
        tk.Label(parent, text=label_text, bg="white", 
                font=("Arial", 9)).grid(row=row, column=0, sticky="w", padx=10, pady=3)
        entry = tk.Entry(parent, width=20)
        entry.insert(0, default_value)
        entry.grid(row=row, column=1, sticky="ew", padx=5, pady=3)
        parent.columnconfigure(1, weight=1)
        setattr(self, attr_name, entry)
    
    def toggle_theme(self):
        self.theme = "dark" if self.theme == "light" else "light"
        self.update_status("Theme toggled (restart for full effect)")
    
    def open_github(self):
        """Open GitHub page in browser"""
        github_url = "https://github.com/aravindhfca"
        try:
            webbrowser.open(github_url)
            self.update_status(f"Opening GitHub profile: {github_url}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open GitHub:\n{str(e)}")
            self.log_error(f"GitHub open error: {str(e)}")
    
    def select_files(self):
        filetypes = [
            ("Excel/CSV files", "*.xlsx *.xlsm *.xls *.csv"),
            ("Excel files", "*.xlsx *.xlsm *.xls"),
            ("CSV files", "*.csv")
        ]
        
        if XLSB_SUPPORT:
            filetypes.insert(0, ("All supported", "*.xlsx *.xlsm *.xls *.xlsb *.csv"))
            filetypes.insert(2, ("XLSB files", "*.xlsb"))
        
        filetypes.append(("All files", "*.*"))
        
        files = filedialog.askopenfilenames(
            title="Select Excel/CSV Files",
            filetypes=filetypes
        )
        if files:
            self.filepaths.extend(files)
            self.update_file_list()
            self.update_sheet_dropdown()
            self.update_status(f"Added {len(files)} file(s)")
            self.sheet_names_per_file = {}
    
    def update_file_list(self):
        self.files_listbox.delete(0, tk.END)
        for filepath in self.filepaths:
            filename = Path(filepath).name
            self.files_listbox.insert(tk.END, filename)
    
    def update_sheet_dropdown(self):
        """Populate sheet name dropdown based on selected files"""
        sheet_choices = set()
        self.sheet_names_per_file = {}
        
        for filepath in self.filepaths:
            if filepath.lower().endswith(".csv"):
                continue
            try:
                if filepath.lower().endswith(".xlsb"):
                    if not XLSB_SUPPORT:
                        continue
                    sheets = self.read_xlsb_sheets(filepath)
                    self.sheet_names_per_file[filepath] = sheets
                    sheet_choices.update(sheets)
                else:
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    self.sheet_names_per_file[filepath] = wb.sheetnames
                    sheet_choices.update(wb.sheetnames)
                    wb.close()
            except Exception as e:
                self.log_error(f"Sheet read error in {filepath}: {str(e)}")
        
        sheet_list = sorted(sheet_choices)
        self.sheet_name_combo['values'] = sheet_list
        if sheet_list:
            self.sheet_name_combo.current(0)
    
    def clear_files(self):
        self.filepaths = []
        self.files_listbox.delete(0, tk.END)
        self.clear_preview()
        self.sheet_names_per_file = {}
        self.selected_sheets = []
        self.sheet_name_combo['values'] = []
        self.available_headers = []
        self.headers_listbox.delete(0, tk.END)
        self.update_status("Files cleared")
    
    def remove_selected(self):
        selected = self.files_listbox.curselection()
        for index in reversed(selected):
            del self.filepaths[index]
        self.update_file_list()
        self.update_sheet_dropdown()
        self.update_status("Selected files removed")
    
    def update_status(self, message):
        self.status_bar.config(text=message)
        self.master.update_idletasks()
    
    def clear_preview(self):
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        self.preview_tree["columns"] = ()
        self.row_count_label.config(text="Rows: 0")
    
    def read_xlsb_sheets(self, filepath):
        """Read sheet names from XLSB file"""
        try:
            if not XLSB_SUPPORT:
                return []
            
            wb = open_xlsb(filepath)
            sheets = [sheet.name for sheet in wb.sheets]
            wb.close()
            return sheets
        except Exception as e:
            self.log_error(f"Error reading XLSB sheets from {filepath}: {str(e)}")
            return []
    
    def detect_column_headers(self):
        """Detect and populate column headers from first file"""
        if not self.filepaths:
            messagebox.showerror("Error", "Please select files first.")
            return
        try:
            filepath = self.filepaths[0]
            headers = []
            
            if filepath.lower().endswith(".csv"):
                with open(filepath, newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)
            elif filepath.lower().endswith(".xlsb"):
                if not XLSB_SUPPORT:
                    messagebox.showerror("Error", "XLSB support not installed. Install pyxlsb: pip install pyxlsb")
                    return
                wb = open_xlsb(filepath)
                sheet = wb.sheets[0]
                first_row = next(sheet.iter_rows())
                headers = [str(cell.v) if cell.v else "" for cell in first_row]
                wb.close()
            else:
                wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                sheet = wb.active
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(cell) if cell else "" for cell in first_row]
                wb.close()
            
            if not headers:
                messagebox.showwarning("Warning", "No headers found in first file.")
                return
            
            # Populate headers listbox
            self.available_headers = headers
            self.headers_listbox.delete(0, tk.END)
            for header in headers:
                self.headers_listbox.insert(tk.END, header)
            
            self.update_status(f"Detected {len(headers)} column headers")
        except Exception as e:
            messagebox.showerror("Error", f"Header detection failed:\n{str(e)}")
            self.log_error(f"Header detection error: {str(e)}")
    
    def clear_headers_selection(self):
        """Clear selected headers"""
        self.headers_listbox.selection_clear(0, tk.END)
        self.update_status("Headers selection cleared")
    
    def auto_detect_headers(self):
        """Auto-detect and display headers in preview tree"""
        if not self.filepaths:
            messagebox.showerror("Error", "Please select files first.")
            return
        try:
            filepath = self.filepaths[0]
            headers = []
            
            if filepath.lower().endswith(".csv"):
                with open(filepath, newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)
            elif filepath.lower().endswith(".xlsb"):
                if not XLSB_SUPPORT:
                    messagebox.showerror("Error", "XLSB support not installed.")
                    return
                wb = open_xlsb(filepath)
                sheet = wb.sheets[0]
                first_row = next(sheet.iter_rows())
                headers = [str(cell.v) if cell.v else "" for cell in first_row]
                wb.close()
            else:
                wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                sheet = wb.active
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(cell) if cell else "" for cell in first_row]
                wb.close()
            
            if not headers:
                messagebox.showwarning("Warning", "No headers found.")
                return
            
            self.preview_tree["columns"] = headers
            self.preview_tree["show"] = "tree headings"
            for idx, header in enumerate(headers):
                self.preview_tree.heading(f"#{idx+1}", text=header)
            self.update_status("Headers auto-detected and displayed in preview.")
        except Exception as e:
            messagebox.showerror("Error", f"Header detection failed:\n{str(e)}")
            self.log_error(f"Header detection error: {str(e)}")
    
    def select_sheets_dialog(self):
        if not self.filepaths:
            messagebox.showerror("Error", "Please select files first.")
            return
        self.sheet_names_per_file = {}
        sheet_choices = set()
        
        for filepath in self.filepaths:
            if filepath.lower().endswith(".csv"):
                continue
            try:
                if filepath.lower().endswith(".xlsb"):
                    if not XLSB_SUPPORT:
                        continue
                    sheets = self.read_xlsb_sheets(filepath)
                    self.sheet_names_per_file[filepath] = sheets
                    sheet_choices.update(sheets)
                else:
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    self.sheet_names_per_file[filepath] = wb.sheetnames
                    sheet_choices.update(wb.sheetnames)
                    wb.close()
            except Exception as e:
                self.log_error(f"Sheet read error in {filepath}: {str(e)}")
        
        if not sheet_choices:
            messagebox.showinfo("Info", "No Excel sheets found in selected files.")
            return
        
        sheet_choices = sorted(sheet_choices)
        selected = simpledialog.askstring("Select Sheets/Tabs",
            f"Enter comma-separated sheet/tab names to merge (e.g. GSTR-2B):\nAvailable: {', '.join(sheet_choices)}")
        
        if selected:
            self.selected_sheets = [s.strip() for s in selected.split(",") if s.strip()]
            self.update_status(f"Selected sheets/tabs: {', '.join(self.selected_sheets)}")
        else:
            self.selected_sheets = []
    
    def preview_data_action(self):
        if not self.filepaths:
            messagebox.showerror("Error", "Please select files first.")
            return
        try:
            self.update_status("Loading preview...")
            self.progress.pack(fill=tk.X, pady=(5, 0))
            self.progress["value"] = 0
            merged_data = self.process_files(preview_mode=True, max_rows=100)
            if not merged_data:
                messagebox.showwarning("Warning", "No data found in selected files.")
                return
            self.display_preview(merged_data)
            self.update_status(f"Preview loaded: {len(merged_data)} rows (max 100 shown)")
        except Exception as e:
            messagebox.showerror("Error", f"Preview failed:\n{str(e)}")
            self.update_status("Preview failed")
            self.log_error(f"Preview error: {str(e)}")
        finally:
            self.progress.pack_forget()
    
    def display_preview(self, data):
        self.clear_preview()
        if not data:
            return
        num_cols = len(data[0])
        columns = [f"Col{i+1}" for i in range(num_cols)]
        self.preview_tree["columns"] = columns
        self.preview_tree["show"] = "tree headings"
        self.preview_tree.column("#0", width=50, minwidth=50)
        self.preview_tree.heading("#0", text="Row")
        for col in columns:
            self.preview_tree.column(col, width=120, minwidth=80)
            self.preview_tree.heading(col, text=col)
        for idx, row in enumerate(data, 1):
            values = [str(cell) if cell is not None else "" for cell in row]
            self.preview_tree.insert("", tk.END, text=str(idx), values=values)
        self.row_count_label.config(text=f"Rows: {len(data)}")
    
    def process_files(self, preview_mode=False, max_rows=None):
        sheet_name = self.sheet_name_var.get().strip()
        start_row = int(self.row_entry.get() or 1)
        start_col_letter = self.col_entry.get().strip().upper() or "A"
        start_col = column_index_from_string(start_col_letter)
        header_mode = self.header_var.get().lower().replace(" ", "_")
        skip_empty = self.skip_empty_var.get()
        remove_duplicates = self.remove_duplicates_var.get()
        add_source = self.add_source_var.get()
        sheet_col_pos = self.sheet_col_var.get()
        add_row_id = self.add_row_id_var.get()
        row_id_pos = self.row_id_var.get() if add_row_id else "none"
        
        merged_data = []
        seen_rows = set()
        first_file = True
        total_files = len(self.filepaths)
        row_counter = 0
        
        for file_idx, filepath in enumerate(self.filepaths):
            self.progress["value"] = ((file_idx + 1) / total_files) * 100
            current_sheet_name = Path(filepath).stem
            self.status_bar.config(text=f"Processing: {Path(filepath).name} ({file_idx+1}/{total_files})")
            self.master.update_idletasks()
            
            try:
                if filepath.lower().endswith(".csv"):
                    with open(filepath, newline='', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        rows = list(reader)
                        if header_mode == "skip_first" and rows:
                            rows = rows[1:]
                        elif header_mode == "first_only" and not first_file and rows:
                            rows = rows[1:]
                        for row in rows[start_row-1:]:
                            if skip_empty and all(cell.strip() == "" for cell in row):
                                continue
                            row = list(row)
                            row_counter += 1
                            
                            # Add sheet name column
                            if sheet_col_pos == "first":
                                row = [current_sheet_name] + row
                            elif sheet_col_pos == "last":
                                row = row + [current_sheet_name]
                            
                            # Add row ID
                            if add_row_id and row_id_pos != "none":
                                if row_id_pos == "first":
                                    row = [row_counter] + row
                                elif row_id_pos == "last":
                                    row = row + [row_counter]
                            
                            if add_source:
                                if sheet_col_pos == "first" or row_id_pos == "first":
                                    row.insert(0 if sheet_col_pos != "first" else 1, Path(filepath).name)
                                else:
                                    row.append(Path(filepath).name)
                            
                            if remove_duplicates:
                                row_hash = hash(tuple(str(cell) for cell in row))
                                if row_hash in seen_rows:
                                    continue
                                seen_rows.add(row_hash)
                            merged_data.append(row)
                            if preview_mode and max_rows and len(merged_data) >= max_rows:
                                return merged_data
                
                elif filepath.lower().endswith(".xlsb"):
                    if not XLSB_SUPPORT:
                        messagebox.showwarning("Warning", f"XLSB format not supported.")
                        continue
                    
                    wb = open_xlsb(filepath)
                    sheets = []
                    if self.selected_sheets:
                        sheets = [wb[s] for s in self.selected_sheets if s in [sheet.name for sheet in wb.sheets]]
                    elif sheet_name:
                        try:
                            sheets = [wb[sheet_name]]
                        except:
                            sheets = wb.sheets
                    else:
                        sheets = wb.sheets
                    
                    for sheet in sheets:
                        sheet_data = []
                        for row in sheet.iter_rows():
                            cells = [cell.v for cell in row[start_col-1:]]
                            if skip_empty and all(cell is None or str(cell).strip() == "" for cell in cells):
                                continue
                            sheet_data.append(cells)
                        
                        if header_mode == "skip_first" and sheet_data:
                            sheet_data = sheet_data[1:]
                        elif header_mode == "first_only" and not first_file and sheet_data:
                            sheet_data = sheet_data[1:]
                        
                        for row in sheet_data[start_row-1:]:
                            row = list(row)
                            row_counter += 1
                            
                            # Add sheet name column
                            if sheet_col_pos == "first":
                                row = [sheet.name] + row
                            elif sheet_col_pos == "last":
                                row = row + [sheet.name]
                            
                            # Add row ID
                            if add_row_id and row_id_pos != "none":
                                if row_id_pos == "first":
                                    row = [row_counter] + row
                                elif row_id_pos == "last":
                                    row = row + [row_counter]
                            
                            if add_source:
                                if sheet_col_pos == "first" or row_id_pos == "first":
                                    row.insert(0 if sheet_col_pos != "first" else 1, Path(filepath).name)
                                else:
                                    row.append(Path(filepath).name)
                            
                            if remove_duplicates:
                                row_hash = hash(tuple(str(cell) for cell in row))
                                if row_hash in seen_rows:
                                    continue
                                seen_rows.add(row_hash)
                            merged_data.append(row)
                            if preview_mode and max_rows and len(merged_data) >= max_rows:
                                wb.close()
                                return merged_data
                    wb.close()
                
                else:
                    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                    sheets = []
                    if self.selected_sheets:
                        sheets = [wb[s] for s in self.selected_sheets if s in wb.sheetnames]
                    elif sheet_name and sheet_name in wb.sheetnames:
                        sheets = [wb[sheet_name]]
                    else:
                        sheets = wb.worksheets
                    
                    for sheet in sheets:
                        sheet_data = []
                        for row in sheet.iter_rows(min_row=start_row, min_col=start_col, values_only=True):
                            if skip_empty and all(cell is None or str(cell).strip() == "" for cell in row):
                                continue
                            sheet_data.append(row)
                        
                        if header_mode == "skip_first" and sheet_data:
                            sheet_data = sheet_data[1:]
                        elif header_mode == "first_only" and not first_file and sheet_data:
                            sheet_data = sheet_data[1:]
                        
                        for row in sheet_data:
                            row = list(row)
                            row_counter += 1
                            
                            # Add sheet name column
                            if sheet_col_pos == "first":
                                row = [sheet.title] + row
                            elif sheet_col_pos == "last":
                                row = row + [sheet.title]
                            
                            # Add row ID
                            if add_row_id and row_id_pos != "none":
                                if row_id_pos == "first":
                                    row = [row_counter] + row
                                elif row_id_pos == "last":
                                    row = row + [row_counter]
                            
                            if add_source:
                                if sheet_col_pos == "first" or row_id_pos == "first":
                                    row.insert(0 if sheet_col_pos != "first" else 1, Path(filepath).name)
                                else:
                                    row.append(Path(filepath).name)
                            
                            if remove_duplicates:
                                row_hash = hash(tuple(str(cell) for cell in row))
                                if row_hash in seen_rows:
                                    continue
                                seen_rows.add(row_hash)
                            merged_data.append(row)
                            if preview_mode and max_rows and len(merged_data) >= max_rows:
                                wb.close()
                                return merged_data
                    wb.close()
                
                first_file = False
            
            except Exception as e:
                messagebox.showwarning("Warning", f"Error reading {Path(filepath).name}:\n{str(e)}")
                self.log_error(f"File read error in {filepath}: {str(e)}")
        
        return merged_data
    
    def merge_files(self):
        if not self.filepaths:
            messagebox.showerror("Error", "Please select files first.")
            return
        try:
            self.update_status("Processing files...")
            self.progress.pack(fill=tk.X, pady=(5, 0))
            self.progress["value"] = 0
            merged_data = self.process_files()
            if not merged_data:
                messagebox.showwarning("Warning", "No data to merge.")
                return
            
            export_format = self.format_var.get()
            if export_format == "xlsx":
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")]
                )
            else:
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv")]
                )
            
            if not save_path:
                self.update_status("Export cancelled")
                return
            
            self.update_status("Exporting...")
            if export_format == "xlsx":
                self.export_excel(merged_data, save_path)
            else:
                self.export_csv(merged_data, save_path)
            
            self.save_config()
            self.update_status(f"Success! Exported {len(merged_data)} rows")
            messagebox.showinfo("Success", 
                              f"Data merged successfully!\n\n"
                              f"Rows: {len(merged_data)}\n"
                              f"File: {Path(save_path).name}")
        except Exception as e:
            messagebox.showerror("Error", f"Merge failed:\n{str(e)}")
            self.update_status("Merge failed")
            self.log_error(f"Merge error: {str(e)}")
        finally:
            self.progress.pack_forget()
    
    def export_excel(self, data, save_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MergedData"
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for row_idx, row_data in enumerate(data, 1):
            ws.append(row_data)
            if row_idx == 1:
                for cell in ws[row_idx]:
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Add summary sheet
        summary = wb.create_sheet("Summary")
        summary.append(["Source File", "Rows Imported"])
        for filepath in self.filepaths:
            summary.append([Path(filepath).name, self.count_rows_in_file(filepath)])
        
        # Add metadata sheet
        metadata = wb.create_sheet("Metadata")
        metadata.append(["Field", "Value"])
        metadata.append(["Total Rows", len(data)])
        metadata.append(["Total Files Merged", len(self.filepaths)])
        metadata.append(["Include Row IDs", self.add_row_id_var.get()])
        metadata.append(["Include Sheet Names", self.sheet_col_var.get() != "none"])
        metadata.append(["Row ID Position", self.row_id_var.get() if self.add_row_id_var.get() else "N/A"])
        metadata.append(["Sheet Name Position", self.sheet_col_var.get()])
        
        wb.save(save_path)
    
    def count_rows_in_file(self, filepath):
        try:
            if filepath.lower().endswith(".csv"):
                with open(filepath, newline='', encoding='utf-8') as f:
                    return sum(1 for _ in f)
            
            elif filepath.lower().endswith(".xlsb"):
                if not XLSB_SUPPORT:
                    return "N/A"
                wb = open_xlsb(filepath)
                if self.selected_sheets:
                    sheet = wb[self.selected_sheets[0]]
                else:
                    sheet = wb.sheets[0]
                count = sum(1 for _ in sheet.iter_rows())
                wb.close()
                return count
            
            else:
                wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                if self.selected_sheets:
                    sheet = wb[self.selected_sheets[0]]
                else:
                    sheet = wb.active
                count = sum(1 for _ in sheet.iter_rows())
                wb.close()
                return count
        except Exception as e:
            self.log_error(f"Row count error in {filepath}: {str(e)}")
            return "Error"
    
    def export_csv(self, data, save_path):
        with open(save_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(data)
    
    def log_error(self, message):
        try:
            with open(self.log_file, "a") as f:
                f.write(message + "\n")
        except:
            pass
    
    def load_config(self):
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    self.theme = config.get('theme', 'light')
        except:
            pass
    
    def save_config(self):
        try:
            config = {'theme': self.theme}
            with open(self.config_file, 'w') as f:
                json.dump(config, f)
        except:
            pass

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()